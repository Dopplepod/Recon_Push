from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Any
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .config import SUMMARY_ORDER, OPERATING_EXPENSE_BUCKETS
from .helpers import (
    normalize_code,
    extract_local_coa_code,
    extract_local_coa_desc,
    first_three_digits,
    pick_sheet_name,
    detect_header_row,
    detect_header_row_from_preview,
    standardize_headers,
    line_item_from_bucket,
    canonical_line_item,
    is_csv_file,
    read_csv_flexible,
)
from .mappings import MappingRepository


class ReconError(Exception):
    pass


@dataclass
class ReconResult:
    payload: Dict


class ReconciliationService:
    def __init__(self, base_dir: Path):
        self.base_dir = base_dir
        self.repo = MappingRepository(base_dir)

    def _load_sap_raw(self, file_obj) -> pd.DataFrame:
        aliases = {
            "gl code": "GL Code",
            "sap mapping": "SAP Mapping",
            "amount": "Amount",
            "gl name": "GL Name",
            "sap description": "SAP Description",
        }
        if is_csv_file(file_obj):
            preview = read_csv_flexible(file_obj, header=None, nrows=10, dtype=str)
            header_row = detect_header_row_from_preview(preview, aliases)
            df = read_csv_flexible(file_obj, header=header_row)
        else:
            xl = pd.ExcelFile(file_obj)
            sheet = pick_sheet_name(xl, "TB BFC")
            header_row = detect_header_row(xl, sheet, aliases)
            df = pd.read_excel(xl, sheet_name=sheet, header=header_row)
        df = standardize_headers(df, aliases)
        required = ["GL Code", "SAP Mapping", "Amount"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            raise ReconError(f"SAP raw file is missing required columns: {', '.join(missing)}")
        work = pd.DataFrame({
            "gl_code": df["GL Code"].map(normalize_code),
            "gl_name": df["GL Name"].fillna("").astype(str).str.strip() if "GL Name" in df.columns else "",
            "sap_mapping_raw": df["SAP Mapping"].map(normalize_code),
            "amount": pd.to_numeric(df["Amount"], errors="coerce").fillna(0.0),
        })
        work = work[(work["gl_code"] != "") & (work["sap_mapping_raw"] != "")]
        return work

    def _detect_os_amount_col(self, df: pd.DataFrame):
        if "Amount" in df.columns and pd.to_numeric(df["Amount"], errors="coerce").notna().any():
            return "Amount", []
        warnings = ["OS raw file has no usable Amount column; structure-only mode loaded with zero amounts."]
        return None, warnings

    def _extract_entity_from_os(self, df: pd.DataFrame) -> str:
        if df.empty or len(df.columns) == 0:
            return ""
        first_col = df.columns[0]
        series = df[first_col].dropna().astype(str).map(str.strip)
        for value in series:
            if not value:
                continue
            digits = "".join(ch for ch in value if ch.isdigit())
            if len(digits) >= 4:
                return digits[:4]
            return value
        return ""

    def _load_os_raw(self, file_obj):
        aliases = {
            "local coa": "Local COA",
            "sap coa": "SAP COA",
            "os coa": "OS COA",
            "amount": "Amount",
        }
        if is_csv_file(file_obj):
            preview = read_csv_flexible(file_obj, header=None, nrows=10, dtype=str)
            header_row = detect_header_row_from_preview(preview, aliases)
            df = read_csv_flexible(file_obj, header=header_row)
        else:
            xl = pd.ExcelFile(file_obj)
            sheet = pick_sheet_name(xl, "OS TB")
            df = pd.read_excel(xl, sheet_name=sheet)
        df = standardize_headers(df, aliases)
        if "Local COA" not in df.columns:
            raise ReconError("OS raw file is missing required column: Local COA")
        entity = self._extract_entity_from_os(df)
        amount_col, warnings = self._detect_os_amount_col(df)
        amt = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0) if amount_col else pd.Series([0.0] * len(df))
        work = pd.DataFrame({
            "local_coa_raw": df["Local COA"].fillna(""),
            "local_coa_code": df["Local COA"].map(extract_local_coa_code),
            "local_coa_desc": df["Local COA"].map(extract_local_coa_desc),
            "sap_code": df["SAP COA"].map(normalize_code) if "SAP COA" in df.columns else "",
            "os_coa": df["OS COA"].map(normalize_code) if "OS COA" in df.columns else "",
            "amount": amt,
        })
        work = work[work["local_coa_code"] != ""].copy()
        return work, warnings, entity

    @staticmethod
    def _is_pl_bucket(value: str) -> bool:
        value = str(value or '').strip()
        digits = ''.join(ch for ch in value if ch.isdigit())
        if len(digits) < 3:
            return False
        try:
            return int(digits[:3]) >= 410
        except ValueError:
            return False

    @staticmethod
    def _join_unique_text(series: pd.Series) -> str:
        vals: List[str] = []
        for v in series.fillna("").astype(str):
            v = v.strip()
            if v and v not in vals:
                vals.append(v)
        return " | ".join(vals)

    @staticmethod
    def _safe_number(value) -> float:
        try:
            return float(value)
        except Exception:
            return 0.0

    def _summary_from_buckets(self, bucket_totals: Dict[str, float]) -> Dict[str, float]:
        revenue = bucket_totals.get("410", 0.0)
        opex = sum(bucket_totals.get(code, 0.0) for code in OPERATING_EXPENSE_BUCKETS)
        ebitda = revenue + opex
        da = bucket_totals.get("504", 0.0)
        ebit = ebitda + da
        finance_income = bucket_totals.get("602", 0.0)
        finance_expense = bucket_totals.get("603", 0.0)
        net_finance = finance_income + finance_expense
        share_results = bucket_totals.get("801", 0.0)
        non_operating = bucket_totals.get("601", 0.0)
        exceptional = bucket_totals.get("699", 0.0)
        pbt = ebit + net_finance + share_results + non_operating + exceptional
        income_tax = bucket_totals.get("607", 0.0)
        discontinued = bucket_totals.get("861", 0.0)
        pat = pbt + income_tax + discontinued
        minority = 0.0
        att_own = pat + minority
        return {
            "Revenue": revenue,
            "Operating Expense (Ex-D And A)": opex,
            "EBITDA": ebitda,
            "EBIT": ebit,
            "6020000 - Finance Income": finance_income,
            "6030000 - Finance Expense": finance_expense,
            "Net Finance (Income / Expense)": net_finance,
            "8010000 - Share Of Results Of AJV": share_results,
            "6010000 - Non Operating Gain Loss": non_operating,
            "6990000 - Exceptional Items": exceptional,
            "Profit Or Loss Before Tax": pbt,
            "6070000 - Income Tax Expense": income_tax,
            "8610000 - Profit Or Loss From Discontinued Operation (Net Of Tax)": discontinued,
            "Net Profit Or Loss (PAT)": pat,
            "PL_MI - Minority Interest": minority,
            "Profit Or Loss Attributable To Owners Of The Company": att_own,
        }

    def _build_core(self, sap_file, os_file, entity: str = "") -> Dict[str, Any]:
        warnings: List[str] = []
        sap_raw = self._load_sap_raw(sap_file)
        os_raw, os_warnings, detected_entity = self._load_os_raw(os_file)
        warnings.extend(os_warnings)

        bfc_map = self.repo.load_bfc_to_os()
        hierarchy_map = self.repo.hierarchy_map()

        sap = sap_raw.merge(
            bfc_map[["sap_mapping_raw", "os_level2", "bucket"]].drop_duplicates(),
            on="sap_mapping_raw",
            how="left",
        )
        sap["bucket"] = sap["bucket"].fillna(sap["os_level2"].map(first_three_digits))
        sap["line_item"] = sap["bucket"].map(lambda x: hierarchy_map.get(x, line_item_from_bucket(x))).map(canonical_line_item)
        sap["detail_id"] = sap["line_item"].fillna("") + "_" + sap["gl_code"]

        os_raw["bucket"] = os_raw["os_coa"].map(first_three_digits)
        os_raw["line_item"] = os_raw["bucket"].map(lambda x: hierarchy_map.get(x, line_item_from_bucket(x))).map(canonical_line_item)
        os_raw["detail_id"] = os_raw["line_item"].fillna("") + "_" + os_raw["local_coa_code"]

        valid_hierarchy_buckets = set(hierarchy_map.keys())

        sap_unmapped = sap[sap["line_item"].fillna("").eq("")].copy()
        os_unmapped = os_raw[os_raw["line_item"].fillna("").eq("")].copy()

        sap_not_in_hierarchy = sap[sap["bucket"].fillna("").ne("") & ~sap["bucket"].isin(valid_hierarchy_buckets)].copy()
        os_not_in_hierarchy = os_raw[os_raw["bucket"].fillna("").ne("") & ~os_raw["bucket"].isin(valid_hierarchy_buckets)].copy()

        sap_non_pl = sap[sap["line_item"].fillna("").ne("") & ~sap["bucket"].map(self._is_pl_bucket)].copy()
        os_non_pl = os_raw[os_raw["line_item"].fillna("").ne("") & ~os_raw["bucket"].map(self._is_pl_bucket)].copy()

        sap_visible = sap[
            sap["line_item"].fillna("").ne("")
            & sap["bucket"].map(self._is_pl_bucket)
            & sap["bucket"].isin(valid_hierarchy_buckets)
        ].copy()
        os_visible = os_raw[
            os_raw["line_item"].fillna("").ne("")
            & os_raw["bucket"].map(self._is_pl_bucket)
            & os_raw["bucket"].isin(valid_hierarchy_buckets)
        ].copy()

        sap_detail = (
            sap_visible.groupby(["bucket", "line_item", "detail_id", "gl_code"], as_index=False)
            .agg(
                sap_amount=("amount", "sum"),
                description=("gl_name", self._join_unique_text),
                route=("sap_mapping_raw", self._join_unique_text),
            )
        )
        sap_detail["display_code"] = sap_detail["gl_code"]

        os_detail = (
            os_visible.groupby(["bucket", "line_item", "detail_id", "local_coa_code"], as_index=False)
            .agg(
                os_amount=("amount", "sum"),
                description=("local_coa_desc", self._join_unique_text),
                raw_os_coa=("os_coa", self._join_unique_text),
                sap_code=("sap_code", self._join_unique_text),
            )
        )
        os_detail["display_code"] = os_detail["local_coa_code"]

        detail_compare = sap_detail.merge(
            os_detail[["bucket", "line_item", "detail_id", "os_amount", "display_code"]],
            on=["bucket", "line_item", "detail_id"],
            how="outer",
            suffixes=("_sap", "_os"),
        )
        detail_compare["sap_amount"] = detail_compare["sap_amount"].fillna(0.0)
        detail_compare["os_amount"] = detail_compare["os_amount"].fillna(0.0)
        detail_compare["difference"] = detail_compare["sap_amount"] - detail_compare["os_amount"]
        detail_compare["display_code"] = detail_compare["display_code_sap"].fillna(detail_compare["display_code_os"]).fillna("")

        bucket_compare = (
            detail_compare.groupby(["bucket", "line_item"], as_index=False)
            .agg(sap_amount=("sap_amount", "sum"), os_amount=("os_amount", "sum"))
        )
        bucket_compare["difference"] = bucket_compare["sap_amount"] - bucket_compare["os_amount"]

        bucket_totals_sap = dict(zip(bucket_compare["bucket"], bucket_compare["sap_amount"]))
        bucket_totals_os = dict(zip(bucket_compare["bucket"], bucket_compare["os_amount"]))
        sap_summary = self._summary_from_buckets(bucket_totals_sap)
        os_summary = self._summary_from_buckets(bucket_totals_os)

        summary_rows = []
        classification_note = ""
        net_finance_diff = self._safe_number(sap_summary.get("Net Finance (Income / Expense)")) - self._safe_number(os_summary.get("Net Finance (Income / Expense)"))
        if abs(net_finance_diff) < 0.5 and (
            abs((self._safe_number(sap_summary.get("6020000 - Finance Income")) - self._safe_number(os_summary.get("6020000 - Finance Income"))) +
                (self._safe_number(sap_summary.get("6030000 - Finance Expense")) - self._safe_number(os_summary.get("6030000 - Finance Expense")))) < 0.5
        ):
            classification_note = "Classification difference only: Finance Income and Finance Expense offset at Net Finance."

        for name in SUMMARY_ORDER:
            s = float(sap_summary.get(name, 0.0))
            o = float(os_summary.get(name, 0.0))
            row = {"line_item": name, "sap_bfc": s, "onestream": o, "difference": s - o}
            if name == "Net Finance (Income / Expense)" and classification_note:
                row["note"] = classification_note
            summary_rows.append(row)

        drilldown_groups = []
        for _, bucket_row in bucket_compare.sort_values(["bucket", "line_item"]).iterrows():
            group_df = detail_compare[
                (detail_compare["bucket"] == bucket_row["bucket"]) &
                (detail_compare["line_item"] == bucket_row["line_item"])
            ].copy().sort_values(["display_code", "detail_id"])
            children = []
            for _, r in group_df.iterrows():
                children.append({
                    "display_code": r.get("display_code", ""),
                    "detail_key": r.get("detail_id", ""),
                    "description": "",
                    "currency": "",
                    "sap_bfc": float(r.get("sap_amount", 0.0)),
                    "onestream": float(r.get("os_amount", 0.0)),
                    "difference": float(r.get("difference", 0.0)),
                })
            drilldown_groups.append({
                "bucket": bucket_row["bucket"],
                "line_item": canonical_line_item(bucket_row["line_item"]),
                "sap_bfc": float(bucket_row["sap_amount"]),
                "onestream": float(bucket_row["os_amount"]),
                "difference": float(bucket_row["difference"]),
                "children": children,
            })

        payload = {
            "meta": {
                "sap_rows": int(len(sap_raw)),
                "os_rows": int(len(os_raw)),
                "gl_codes": int(sap_raw["gl_code"].nunique()),
                "unmapped_gl_codes": 0,
                "entity": detected_entity or entity or "-",
            },
            "summary": summary_rows,
            "drilldown": drilldown_groups,
            "debug": {
                "warnings": warnings,
                "excluded_non_pl": {
                    "sap_rows": int(len(sap_non_pl)),
                    "os_rows": int(len(os_non_pl)),
                    "sap_line_items": sorted(sap_non_pl["line_item"].dropna().astype(str).unique().tolist())[:100],
                    "os_line_items": sorted(os_non_pl["line_item"].dropna().astype(str).unique().tolist())[:100],
                    "rule": "Only P&L buckets with first 3 digits >= 410 and present in the hierarchy are included in visible summary and drilldown.",
                },
                "excluded_not_in_hierarchy": {
                    "sap_rows": int(len(sap_not_in_hierarchy)),
                    "os_rows": int(len(os_not_in_hierarchy)),
                    "sap_buckets": sorted(sap_not_in_hierarchy["bucket"].dropna().astype(str).unique().tolist())[:100],
                    "os_buckets": sorted(os_not_in_hierarchy["bucket"].dropna().astype(str).unique().tolist())[:100],
                    "rule": "Buckets not defined in hierarchy.xml are excluded from visible outputs to reduce noise.",
                },
                "methodology": {
                    "sap_id": "OS Level 2 + '_' + GL Code",
                    "sap_join_key": "Raw SAP Mapping with suffix (exact match)",
                    "os_id": "OS Level 2 + '_' + Local COA code",
                    "os_level2_basis": "derived from first 3 digits of raw OS COA and expanded to full label",
                    "entity_model": "Entity-agnostic. The backend logic is reusable across entities as long as the uploaded raw files follow the same structure.",
                },
                "unmapped_sap": [],
                "unmapped_os": [],
            }
        }

        gl_compare = (
            sap_raw.groupby("gl_code", as_index=False).agg(sap_bfc=("amount", "sum"))
            .merge(os_raw.groupby("local_coa_code", as_index=False).agg(onestream=("amount", "sum")),
                   left_on="gl_code", right_on="local_coa_code", how="outer")
        )
        gl_compare["gl_code"] = gl_compare["gl_code"].fillna(gl_compare["local_coa_code"]).fillna("")
        gl_compare["sap_bfc"] = gl_compare["sap_bfc"].fillna(0.0)
        gl_compare["onestream"] = gl_compare["onestream"].fillna(0.0)
        gl_compare["difference"] = gl_compare["sap_bfc"] - gl_compare["onestream"]
        gl_compare = gl_compare[["gl_code", "sap_bfc", "onestream", "difference"]]
        gl_compare = gl_compare[gl_compare["difference"].abs() > 1e-9].sort_values("gl_code")

        return {
            "payload": payload,
            "sap_raw": sap_raw,
            "os_raw": os_raw,
            "gl_compare": gl_compare,
            "detected_entity": detected_entity or entity or "-",
        }

    def reconcile(self, sap_file, os_file, entity: str = "") -> ReconResult:
        return ReconResult(self._build_core(sap_file, os_file, entity)["payload"])

    @staticmethod
    def _apply_number_format(cell):
        cell.number_format = '#,##0;[Red](#,##0);-'

    def _style_table_sheet(self, ws, headers, *, diff_cols=None, header_row=2, freeze=None, auto_filter=True):
        diff_cols = set(diff_cols or [])
        title_fill = PatternFill('solid', fgColor='DCE6F1')
        header_fill = PatternFill('solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        section_fill = PatternFill('solid', fgColor='EEF5FB')
        yellow_fill = PatternFill('solid', fgColor='FFF59D')
        thin_gray = Side(style='thin', color='D9E2F3')

        ws.sheet_view.showGridLines = False
        if freeze:
            ws.freeze_panes = freeze

        for row in range(1, header_row):
            for cell in ws[row]:
                cell.fill = title_fill
                cell.font = Font(bold=True, color='1F1F1F')
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(bottom=thin_gray)

        for row in ws.iter_rows(min_row=header_row + 1):
            row_has_diff = False
            row_type = str(row[0].value or '').strip().lower() if len(row) else ''
            for col_idx, cell in enumerate(row, start=1):
                if isinstance(cell.value, (int, float)):
                    self._apply_number_format(cell)
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
                if col_idx in diff_cols and isinstance(cell.value, (int, float)) and abs(cell.value) > 1e-9:
                    row_has_diff = True
            if row_type in {'group', 'summary'}:
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = section_fill
            if row_has_diff:
                for col_idx in diff_cols:
                    row[col_idx - 1].fill = yellow_fill

        if auto_filter and ws.max_row >= header_row:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells[:3000]:
                val = '' if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            width = min(max(max_len + 2, 12), 42)
            ws.column_dimensions[col_letter].width = width

    def _write_sheet_title(self, ws, title, subtitle, columns):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color='1F1F1F')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
        ws['A2'] = subtitle
        ws['A2'].font = Font(italic=True, size=10, color='666666')
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

    def export_excel(self, sap_file, os_file, entity: str = "") -> BytesIO:
        core = self._build_core(sap_file, os_file, entity)
        payload = core['payload']
        wb = Workbook()
        wb.remove(wb.active)

        entity_label = core['detected_entity']

        # Summary
        ws = wb.create_sheet('Summary', 0)
        summary_headers = ['Line Item', 'SAP BFC', 'OneStream', 'Difference', 'Note']
        self._write_sheet_title(ws, 'Reconciliation Summary', f'Entity: {entity_label}', len(summary_headers))
        self._style_table_sheet(ws, summary_headers, diff_cols={4}, header_row=3, freeze='A4')
        for row in payload['summary']:
            ws.append([row.get('line_item', ''), row.get('sap_bfc', 0.0), row.get('onestream', 0.0), row.get('difference', 0.0), row.get('note', '')])

        # Drilldown expanded
        ws = wb.create_sheet('Drilldown Expanded', 1)
        drill_headers = ['Type', 'Line Item', 'GL Code', 'Description', 'Currency', 'SAP BFC', 'OneStream', 'Difference']
        self._write_sheet_title(ws, 'Expanded Drilldown', 'Fully expanded drilldown by visible hierarchy buckets only', len(drill_headers))
        self._style_table_sheet(ws, drill_headers, diff_cols={8}, header_row=3, freeze='A4')
        for group in payload['drilldown']:
            ws.append(['Group', group.get('line_item', ''), '', '', '', group.get('sap_bfc', 0.0), group.get('onestream', 0.0), group.get('difference', 0.0)])
            for child in group.get('children', []):
                ws.append(['Detail', group.get('line_item', ''), child.get('display_code', ''), child.get('description', ''), child.get('currency', ''), child.get('sap_bfc', 0.0), child.get('onestream', 0.0), child.get('difference', 0.0)])

        # GL Compare differences only
        ws = wb.create_sheet('GL Compare Diff', 2)
        compare_headers = ['GL Code', 'SAP BFC', 'OneStream', 'Difference']
        self._write_sheet_title(ws, 'GL Code Differences Only', 'Side-by-side comparison for mismatched GL codes only', len(compare_headers))
        self._style_table_sheet(ws, compare_headers, diff_cols={4}, header_row=3, freeze='A4')
        for _, row in core['gl_compare'].iterrows():
            ws.append([row['gl_code'], float(row['sap_bfc']), float(row['onestream']), float(row['difference'])])

        # SAP Raw normalized
        ws = wb.create_sheet('SAP Raw', 3)
        sap_headers = ['GL Code', 'GL Name', 'SAP Mapping', 'Amount']
        self._write_sheet_title(ws, 'SAP Raw Data', 'Normalized SAP raw upload used in reconciliation', len(sap_headers))
        self._style_table_sheet(ws, sap_headers, header_row=3, freeze='A4')
        for _, row in core['sap_raw'].iterrows():
            ws.append([row.get('gl_code', ''), row.get('gl_name', ''), row.get('sap_mapping_raw', ''), float(row.get('amount', 0.0))])

        # OS Raw normalized
        ws = wb.create_sheet('OS Raw', 4)
        os_headers = ['Local COA Raw', 'Local COA Code', 'Description', 'SAP COA', 'OS COA', 'Amount']
        self._write_sheet_title(ws, 'OS Raw Data', 'Normalized OneStream raw upload used in reconciliation', len(os_headers))
        self._style_table_sheet(ws, os_headers, header_row=3, freeze='A4')
        for _, row in core['os_raw'].iterrows():
            ws.append([row.get('local_coa_raw', ''), row.get('local_coa_code', ''), row.get('local_coa_desc', ''), row.get('sap_code', ''), row.get('os_coa', ''), float(row.get('amount', 0.0))])

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.sheet_properties.tabColor = '1F4E78'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
